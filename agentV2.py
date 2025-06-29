import os
import re
import json
import pandas as pd
import time
from groq import Groq
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

# Inicializa cliente Groq usando variable de entorno
client = Groq(api_key=os.getenv("GROQ_API_KEY"))

# Estructura de campañas: clave -> (número de items, límite de caracteres)
CAMPAIGNS_STRUCTURE = {
    "SEM": {
        "headlines": (15, 30),
        "long_headlines": (5, 90),
        "short_description": (1, 60),
        "long_descriptions": (4, 90)
    },
    "MetaDemandGen": {
        "primary_texts": (4, 250),
        "headlines": (3, 30),
        "descriptions": (3, 30)
    },
    "MetaDemandCapture": {
        "primary_texts": (4, 250),
        "headlines": (3, 30),
        "descriptions": (3, 30)
    },
    "GoogleDemandGen": {
        "headlines": (5, 30),
        "short_description": (1, 60),
        "long_descriptions": (4, 90)
    },
    "GooglePMAX": {
        "headlines": (15, 30),
        "long_headlines": (5, 90),
        "short_description": (1, 60),
        "long_descriptions": (4, 90)
    }
}

# --------------------------- Funciones Auxiliares ---------------------------
def cargar_referencias(path: str) -> pd.DataFrame:
    return pd.read_excel(path, sheet_name="Copies")

def cargar_contenidos(path: str) -> pd.DataFrame:
    return pd.read_excel(path)

def cargar_planes(path: str) -> pd.DataFrame:
    return pd.read_excel(path)

def cargar_specs(path: str) -> pd.DataFrame:
    return pd.read_excel(path)

def obtener_info_contenido(brief: str,
                           content_df: pd.DataFrame,
                           plans_df: pd.DataFrame) -> tuple:
    brief_lower = brief.lower()

    mask_nombre = content_df["content_name"].astype(str).str.lower().apply(lambda x: x in brief_lower)
    filas_match = content_df[mask_nombre]
    
    if filas_match.empty:
        mask_pais = content_df["content_country"].astype(str).str.lower().apply(lambda x: x in brief_lower)
        filas_match = content_df[mask_pais]

    if filas_match.empty:
        filas_match = content_df.head(1)
        
    fila = filas_match.iloc[0]
    content_info = {
        "content_name": fila["content_name"],
        "languages": [lang.strip() for lang in str(fila.get("content_languages", "")).split(",") if lang.strip()],
        "details": fila.get("content_details", ""),
        "markets": [m.strip() for m in str(fila.get("markets_available", "")).split(",") if m.strip()]
    }
    
    plans_available = [p.strip() for p in str(fila.get("plans_available", "")).split(",") if p.strip()]
    plan_info = {}
    
    for market in content_info["markets"]:
        planes = []
        for plan_nom in plans_available:
            
            mask_plan = (
                (plans_df["plan_name"].astype(str).str.lower() == plan_nom.lower()) &
                (plans_df["markets"].astype(str).str.upper().str.contains(market.upper(), na=False))
            )
            
            filas_plan = plans_df[mask_plan]
            if not filas_plan.empty:
                p = filas_plan.iloc[0]
                planes.append({
                    "plan_name": p["plan_name"],
                    "recurring_period": p["recurring_period"],
                    "price": p["price"],
                    "currency": p["currency"],
                    "currency_symbol": p["currency_symbol"],
                    "has_discount": p["has_discount"],
                    "marketing_discount": p["marketing_discount"]
                })
        plan_info[market] = planes
        
    return content_info, plan_info

def limpiar_json(response_text: str) -> dict:
    if not response_text:
        print("ADVERTENCIA: La respuesta de la API estaba vacía.")
        return {}
        
    start = response_text.find('{')
    end = response_text.rfind('}') + 1
    
    if start == -1:
        print("ADVERTENCIA: La respuesta de la API no contenía un objeto JSON.")
        return {}

    raw = response_text[start:end]
    try:
        return json.loads(raw, strict=False)
    except json.JSONDecodeError:
        print("ADVERTENCIA: No se pudo decodificar el JSON extraído de la respuesta.")
        return {}

def preparar_batch(texts: list, limit: int, tipo: str) -> list:
    df = pd.DataFrame({"Original": texts, "Reescrito": texts.copy()})
    over_limit = df[df["Original"].str.len() > limit].index.tolist()
    if over_limit:
        bloques = "\n".join([
            f'Texto {i+1}: "{df.at[i, "Original"]}"' for i in over_limit
        ])
        prompt = f"""
Eres un redactor publicitario experto.
Reescribe los siguientes textos para que no superen {limit} caracteres cada uno.
Mantén el sentido original y estilo.
Devuelve sólo los textos reescritos, uno por línea, sin numeración ni comillas.

{bloques}
""".strip()
        resp = client.chat.completions.create(
            messages=[{"role": "system", "content": "You are a helpful assistant."},
                      {"role": "user", "content": prompt}],
            model="llama-3.3-70b-versatile",
            temperature=0.3
        )
        lines = [l.strip() for l in resp.choices[0].message.content.splitlines() if l.strip()]
        for idx, new_text in zip(over_limit, lines):
            clean = new_text if len(new_text) <= limit else new_text[:limit].rstrip() + "..."
            df.at[idx, "Reescrito"] = clean
    return df["Reescrito"].tolist()

def traducir_batch(texts: list, target: str) -> list:
    if target not in ("en", "pt"):
        return texts
    lenguaje = "English (US)" if target == "en" else "Português (Brasil)"
    input_block = json.dumps(texts, ensure_ascii=False, indent=2)
    prompt = f"""
Eres un traductor experto. A continuación tienes una lista de textos en español, 
cada uno puede contener varias líneas, bullets y saltos de línea.
Traduce cada texto completo al {lenguaje}, preservando exactamente su estructura interna
(líneas, bullets y saltos de línea).

Devuelve SÓLO un JSON con un campo "translations" que sea un array de strings,
en el mismo orden de los textos de entrada.

Textos a traducir:
{input_block}
""".strip()
    resp = client.chat.completions.create(
        messages=[{"role": "system", "content": "You are a helpful assistant."},
                  {"role": "user", "content": prompt}],
        model="llama-3.3-70b-versatile",
        temperature=0
    )
    content = resp.choices[0].message.content
    match = re.search(r'(\{.*\})', content, flags=re.S)
    if match:
        try:
            data = json.loads(match.group(1))
            translations = data.get("translations", [])
            return [translations[i] if i < len(translations) else texts[i] for i in range(len(texts))]
        except json.JSONDecodeError:
            pass
    return texts

def generar_prompt_multi(briefs: dict,
                         ref_df: pd.DataFrame,
                         content_info: dict,
                         plan_info: dict,
                         specs_df: pd.DataFrame) -> str:
                             
    # 1. Construir ejemplos con contexto de mercado
    ejemplos = []
    muestra = ref_df.sample(n=min(10, len(ref_df)), random_state=42)
    for _, row in muestra.iterrows():
        market_info = row.get('Market', 'General')
        ejemplos.append(
            f"- [{market_info}] [{row['Idioma']}] {row['Platform']} {row['Tipo']} {row['Campo']}: \"{row['Texto']}\""
        )
    ejemplos_block = "\n".join(ejemplos)

    # 2. Construir la estructura JSON de salida
    template = {}
    for market in content_info["markets"]:
        template[market] = {}
        for camp, fields in CAMPAIGNS_STRUCTURE.items():
            template[market][camp] = {field: ["" for _ in range(count)] if count > 1 else "" for field, (count, _) in fields.items()}

    # 3. Construir bloque de información factual
    info_block = [f"Contenido identificado: {content_info['content_name']}"]
    if content_info.get("details"): info_block.append(f"Detalles: {content_info['details']}")
    info_block.append(f"Idiomas de narración disponibles: {', '.join(content_info['languages'])}")
    info_block.append("")
    info_block.append("Mercados y planes disponibles (Fuente de Verdad para Precios):")
    for market, planes in plan_info.items():
        planes_str = "; ".join([
            f"{p['plan_name']} ({p['recurring_period']}) a {p['currency_symbol']}{p['price']} {p['currency']}" + (f" con {p['marketing_discount']}% de descuento" if p['has_discount'].lower()=="si" else "")
            for p in planes
        ]) or "Sin planes configurados"
        info_block.append(f"- Mercado {market}: {planes_str}")

    # 4. Definimos la nueva personalidad y las instrucciones de alto rendimiento
    personalidad = (
        "Eres 'Wilson', el redactor publicitario #1 del mundo especializado en marketing deportivo y servicios de streaming por suscripción. "
        "Tu estilo es apasionado, enérgico y siempre centrado en la emoción del hincha. Entiendes la urgencia y la emoción del fútbol en vivo. "
        "Tu misión es crear copies que generen una necesidad irresistible de ver el partido AHORA."
    )

    instrucciones_maestras = [
        "Proceso de Pensamiento Obligatorio: Antes de escribir cada texto, debes seguir mentalmente estos 3 pasos: 1. ¿Cuál es el beneficio más potente para el hincha (ej. exclusividad, ver a su equipo, no perderse el clásico)? 2. ¿Qué emoción quiero provocar (urgencia, pasión, miedo a perdérselo)? 3. ¿Cuál es el llamado a la acción más directo? Una vez que tengas esto claro, escribe el copy.",
        "Regla de Oro de Calidad y Longitud: Es INACEPTABLE entregar textos cortos, genéricos o que no llenen el espacio. Tu reputación como el mejor depende de ello. DEBES maximizar el uso del espacio disponible, apuntando siempre al 95-100% del límite de caracteres. Cada copy debe ser vibrante, específico y rebosar energía.",
        "Regla de Datos: Para precios, monedas o descuentos, la sección 'Mercados y planes disponibles' es tu única fuente de verdad y su uso es obligatorio."
    ]

    especificaciones_por_campaña = ["\nEspecificaciones detalladas por tipo de campaña:"]
    for _, row in specs_df.iterrows():
        campo_limpio = row['title'].split(' ')[0].lower().replace(' ','_')
        instrucciones_maestras.append(
            f"- Para la campaña '{row['campaign']}' en el campo '{campo_limpio}':"
            f" Estilo requerido: {row['style']}. Detalles importantes: {row['details']}. Objetivo de comunicación: {row['objective']}."
        )

    # 5. Ensamblar el prompt final con la nueva estructura
    prompt = (
        f"{personalidad}\n\n"
        f"=== CONTEXTO GENERAL ===\n"
        f"Empresa: {briefs['company']}\n"
        f"Nombre de campaña: {briefs['campaign_name']}\n"
        f"Brief: {briefs['campaign_brief']}\n\n"
        f"=== DATOS DEL PRODUCTO Y PLANES (Fuente de Verdad) ===\n"
        f"{'\n'.join(info_block)}\n\n"
        f"=== EJEMPLOS DE ESTILO Y TONO ===\n"
        f"{ejemplos_block}\n\n"
        f"=== TUS REGLAS Y PROCESO MENTAL ===\n"
        f"{'\n'.join(instrucciones_maestras)}\n\n"
         f"=== TAREA FINAL: DEVOLVER SÓLO EL JSON ===\n"
        f"Tu única y exclusiva salida debe ser el siguiente objeto JSON, completado en su totalidad. No incluyas '```json', saludos, ni ningún otro texto fuera del objeto JSON.\n"
        f"{json.dumps(template, ensure_ascii=False, indent=2)}"
    )
    return prompt

def generar_excel_multi(data: dict, filename: str = "copies_final.xlsx"):
    rows = []
    for market, market_data in data.items():
        for camp, fields in CAMPAIGNS_STRUCTURE.items():
            platform_map = {"SEM":"Google","GoogleDemandGen":"Google","GooglePMAX":"Google","MetaDemandGen":"Meta","MetaDemandCapture":"Meta"}
            platform = platform_map.get(camp,"")
            tipo_map = {"SEM":"SEM","GoogleDemandGen":"DemandGen","GooglePMAX":"PMAX","MetaDemandGen":"DemandGen","MetaDemandCapture":"DemandCapture"}
            tipo = tipo_map.get(camp, camp)
            camp_data = market_data.get(camp, {})
            for field,(count,limit) in fields.items():
                orig = camp_data.get(field, [])
                orig_list = orig if isinstance(orig,list) else [orig]
                es_clean = preparar_batch(orig_list, limit, f"{camp}.{field}")
                en_clean = preparar_batch(traducir_batch(es_clean, 'en'), limit, f"{camp}.{field}.en")
                pt_clean = preparar_batch(traducir_batch(es_clean, 'pt'), limit, f"{camp}.{field}.pt")
                for lst in (es_clean,en_clean,pt_clean):
                    while len(lst)<count: lst.append("")
                for i in range(count):
                    title = f"{field.replace('_',' ').title()} {i+1}"
                    for lang, texts in (('es',es_clean),('en',en_clean),('pt',pt_clean)):
                        txt = texts[i]
                        rows.append({
                            "Market":market,
                            "Platform":platform,
                            "Tipo":tipo,
                            "Campo":field,
                            "Título":title,
                            "Idioma":lang,
                            "Texto":txt,
                            "Caracteres":len(txt),
                            "Max Caracteres":limit,
                            "Check":1 if len(txt)<=limit else 0
                        })
                print(f"Pausa de 10 segundos después de procesar el campo: {camp}/{field}")
                time.sleep(10)
    df = pd.DataFrame(rows)
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Copies", index=False)
    wb = load_workbook(filename)
    ws = wb["Copies"]
    for r in range(2, ws.max_row+1):
        if isinstance(ws.cell(r,7).value,str) and ws.cell(r,7).value.endswith("..."):
            for c in range(1,7):
                cell = ws.cell(r,c)
                cell.font = Font(color="9C0006")
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    wb.save(filename)

def generar_copies(campaign_name: str, campaign_brief: str, output_filename: str = "copies_generadas.xlsx") -> str:
    briefs_config = {
        "company": os.getenv("COMPANY_NAME","Fanatiz"),
        "company_context": os.getenv("COMPANY_CONTEXT",
            "Empresa pionera en transmitir fútbol de sudamérica y el mundo fuera de sudamérica, "
            "principalmente en Estados Unidos, España y Canadá, con sus contenidos principales como la Liga Argentina, "
            "la Liga1 de Perú, la Primera División de Paraguay, la Liga BetPlay de Colombia, la Primera División de Uruguay, "
            "Copa Libertadores y Sudamericana, La LigaPro de Ecuador, la Primeira Liga de Portugal, la Ligue 1 de Francia, "
            "la SuperLig de Turquía, la CAF Champions League, el Premier Padel Tour y mucho más. Fanatiz transmite el fútbol "
            "de manera 100% legal y seguro, ofreciendo el contenido en alta calidad, con el idioma de preferencia "
            "(Español, Inglés o Portugués) o con la opción de seleccionar el sonido del estadio para vivir los partidos "
            "como si estuvieras allí. Pueden usar la app desde nuestro navegador web, descargarla en celular Android o Apple "
            "Store con la posibilidad de castear a la TV o si tienen Smart TV, Android TV, Apple TV, Roku, FireTV, Samsung, "
            "LG o TV Boxes pueden descargar la aplicación directamente en su TV y disfrutar el fútbol en la pantalla grande "
            "(se recomienda usar el navegador de la TV)"),
        "value_proposition": os.getenv("VALUE_PROPOSITION",
            "Fútbol 100% legal y seguro, en tu idioma, en vivo, alta calidad y en el dispositivo de tu preferencia: "
            "TV descargando la app de Fanatiz, casteando desde el celular o computadora o usando el navegador de la TV, "
            "computadora o celular"),
        "campaign_name": campaign_name,
        "campaign_brief": campaign_brief,
        "extras": os.getenv("CAMPAIGN_EXTRAS","")
    }

    # Forma correcta de obtener la ruta de la carpeta del proyecto
    base_path = os.path.abspath(os.path.dirname(__file__))
    df_refs = cargar_referencias(os.path.join(base_path, "Mejor_Performing_Copies_Paid_Fanatiz.xlsx"))
    df_content = cargar_contenidos(os.path.join(base_path, "content_by_country.xlsx"))
    df_plans = cargar_planes(os.path.join(base_path, "plans_and_pricing.xlsx"))
    df_specs = cargar_specs(os.path.join(base_path, "platforms_and_campaigns_specs.xlsx"))
    content_info, plan_info = obtener_info_contenido(briefs_config["campaign_brief"], df_content, df_plans)
    prompt = generar_prompt_multi(briefs_config, df_refs, content_info, plan_info, df_specs)
    resp = client.chat.completions.create(
        messages=[
            {
                "role": "system",
                "content": "Eres un asistente experto en devolver respuestas estructuradas. Tu única función es generar un objeto JSON basado en las instrucciones del usuario. No debes añadir texto introductorio, explicaciones, ni comentarios. Tu salida debe ser exclusivamente el código JSON."
            },
            {
                "role": "user",
                "content": prompt
            }
        ],
        model="llama-3.3-70b-versatile", 
        temperature=0.35, # Subimos un poco la temperatura para fomentar la creatividad de 'Wilson'
    )
    data = limpiar_json(resp.choices[0].message.content)
    generar_excel_multi(data, filename=output_filename)
    return output_filename
