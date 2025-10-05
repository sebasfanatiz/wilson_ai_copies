# agente_opt.py
# ==============================================================================
# 1) IMPORTS
# ==============================================================================
import os
import re
import json
import time
import random
import sys
import copy
import math
import pandas as pd
import unicodedata
from openai import OpenAI
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

# ==============================================================================
# 2) CONFIG & CONSTANTS
# ==============================================================================
MODEL_CHAT = os.getenv("OPENAI_MODEL_CHAT", "gpt-5-mini")
client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

# Estructura por campa√±a (conteos/limites)
CAMPAIGNS_STRUCTURE = {
    "SEM": {"headlines": (14, 30), "long_headlines": (4, 90)},
    "MetaDemandGen": {"primary_texts": (4, 250), "headlines": (3, 30), "descriptions": (3, 30)},
    "MetaDemandCapture": {"primary_texts": (4, 250), "headlines": (3, 30), "descriptions": (3, 30)},
    "GoogleDemandGen": {"headlines": (5, 30), "short_description": (1, 60), "long_descriptions": (4, 90)},
    "GooglePMAX": {"headlines": (15, 30), "long_headlines": (5, 90), "short_description": (1, 60), "long_descriptions": (4, 90)},
}

# M√≠nimos
MIN_CHARS_BY_FIELD = {"primary_texts": 200}

# Opcional (solo para legibilidad/seguridad adicional):
EXPANDABLE_FIELDS = {"primary_texts"}  # whitelist

# Defaults de planes por plataforma
DEFAULT_PLANS_BY_PLATFORM = {
    "Fanatiz": ["Front Row Monthly", "Front Row Annual"],
    "L1MAX": ["L1MAX Full", "L1MAX M√≥vil", "L1MAX Full Anual"],
    "AFA Play": ["AFA Play", "AFA Play Annual"],
}

COMPANY_BY_PLATFORM = {"Fanatiz": "Fanatiz", "L1MAX": "L1MAX", "AFA Play": "AFA Play"}

STOPWORDS_END_ES = {
    "de","del","en","con","para","por","vs","y","o","a","al","la","el","los","las","un","una","unos","unas",
    "lo","su","sus","tu","tus","mi","mis","este","esta","estos","estas","ese","esa","esos","esas"
}

# ==============================================================================
# 3) UTILITIES
# ==============================================================================

def _normalize_headers(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip().lower().replace("\n", " ").replace("  ", " ").replace(" ", "_") for c in df.columns]
    return df


def _ensure_columns(df: pd.DataFrame, required_cols: list, df_name: str = "DataFrame"):
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise KeyError(f"{df_name} no contiene las columnas requeridas: {missing}. Columns disponibles: {list(df.columns)}")


def _parse_langs(s: str) -> tuple:
    m = {"es", "en", "pt"}
    if not s:
        return ("es",)
    items = [t.strip().lower() for t in s.split(",") if t.strip()]
    items = [t for t in items if t in m]
    return tuple(sorted(set(items), key=items.index)) or ("es",)


def _norm_base_name(plan_name: str) -> str:
    if not isinstance(plan_name, str):
        return ""
    return plan_name.lower().replace("monthly", "").replace("annual", "").replace("anual", "").replace("mensual", "").strip()


def _split_markets_cell(cell: str) -> list[str]:
    if not isinstance(cell, str):
        return []
    return [t.strip() for t in cell.split(',') if t.strip()]


def _seems_english(txt: str) -> bool:
    if not isinstance(txt, str):
        return False
    t = txt.lower()
    common = (" the ", " to ", " and ", " for ", " with ", " your ", " watch ")
    return sum(w in f" {t} " for w in common) >= 2

def _norm_str(s: str) -> str:
    if not isinstance(s, str):
        return ""
    s = s.replace("\u00A0", " ")  # NBSP -> espacio normal
    s = unicodedata.normalize("NFKD", s)  # separa diacr√≠ticos
    s = "".join(ch for ch in s if not unicodedata.combining(ch))  # quita acentos
    s = s.strip().lower()
    s = re.sub(r"\s+", " ", s)  # colapsa espacios
    return s

def _polish_ending_es(t: str) -> str:
    """Quita ‚Äúcolas‚Äù feas (conectores sueltos, dos puntos, guiones, 'on' recortado) y deja la frase cerrada."""
    if not isinstance(t, str):
        return ""
    s = t.strip()

    # casitos t√≠picos de truncado feo
    s = re.sub(r"[:‚Äî\-‚Äì]\s*$", "", s)          # termina en ':' o gui√≥n
    s = re.sub(r"\s+(?:vs|VS)\s*$", "", s)     # queda "Navegantes vs"
    s = re.sub(r"\s+(?:o|y)\s+on\s*$", "", s)  # "en vivo o on" -> corta ' o on'
    s = re.sub(r"\s+on\s*$", "", s)            # " ... on" (origen: 'on demand' cortado)
    s = re.sub(r"\s+(?:en|de|del)\s*$", "", s) # " ... en", " ... de"

    # si termina con stopword, vamos quitando la √∫ltima palabra hasta que cierre bien
    tokens = s.split()
    while tokens and tokens[-1].lower() in STOPWORDS_END_ES:
        tokens.pop()
    s = " ".join(tokens).strip()

    # si qued√≥ vac√≠o por limpieza agresiva, volvemos al original sin la √∫ltima palabra
    if not s and t.strip():
        base = t.strip().rsplit(" ", 1)[0] if " " in t.strip() else t.strip()
        s = base.strip(":‚Äî-‚Äì ").strip()

    return s

def _smart_trim(text: str, limit: int, lang: str = "es") -> str:
    """Recorta con preferencia por cierre de frase; nunca corta palabra ni deja conectores sueltos."""
    if not isinstance(text, str):
        return ""
    t = text.strip()
    if len(t) <= limit:
        return _polish_ending_es(t) if lang == "es" else t

    # 1) intentar cortar en puntuaci√≥n ‚Äúfina‚Äù
    cut_points = [m.end() for m in re.finditer(r"[\.!?‚Ä¶]\s", t)]
    cut_points = [c for c in cut_points if c <= limit]
    if cut_points:
        return _polish_ending_es(t[:max(cut_points)].strip()) if lang == "es" else t[:max(cut_points)].strip()

    # 2) cortar por √∫ltimo espacio ‚Äúsaludable‚Äù
    last_space = t.rfind(" ", 0, limit + 1)
    if last_space != -1 and last_space >= int(limit * 0.6):
        return _polish_ending_es(t[:last_space].strip()) if lang == "es" else t[:last_space].strip()

    # 3) ensamblar palabras completas hasta el l√≠mite
    out = []
    for w in t.split():
        if len((" ".join(out + [w])).strip()) <= limit:
            out.append(w)
        else:
            break
    s = " ".join(out).strip()
    return _polish_ending_es(s) if lang == "es" else s

def _format_discount(val) -> str:
    s = str(val).strip()
    if not s:
        return ""
    # 1) ya trae el s√≠mbolo
    if s.endswith("%"):
        try:
            float(s[:-1].replace(",", "."))
            return s
        except:
            return s
    # 2) decimal tipo 0.25
    try:
        f = float(s.replace(",", "."))
        if 0 < f < 1:
            return f"{int(round(f * 100))}%"
    except:
        pass
    # 3) n√∫mero entero "25"
    if re.fullmatch(r"\d{1,3}", s):
        return f"{s}%"
    # 4) extraer un porcentaje si est√° embebido
    m = re.search(r"(\d{1,3})\s*%", s)
    if m:
        return f"{m.group(1)}%"
    return s  # fallback tal cual



# ==============================================================================
# 4) DATA LOADERS
# ==============================================================================

def cargar_referencias(path): return pd.read_excel(path, sheet_name="Copies")

def cargar_contenidos(path):  return pd.read_excel(path)

def cargar_planes(path):      return pd.read_excel(path)

def cargar_specs(path):       return pd.read_excel(path)


# ==============================================================================
# 5) CONTENT LOOKUP / FILTERS
# ==============================================================================

def get_platform_plans_set(df_plans, platform):
    if 'platform' not in df_plans.columns: raise KeyError("'platform' no existe en df_plans")
    if 'plan_name' not in df_plans.columns: raise KeyError("'plan_name' no existe en df_plans")
    mask = df_plans['platform'].fillna("").str.lower() == platform.lower()
    dfp = df_plans[mask]
    base_names = {_norm_base_name(n) for n in dfp['plan_name'].dropna().astype(str)}
    full_names = set(dfp['plan_name'].dropna().astype(str))
    return base_names, full_names


def any_plan_matches_platform(plans_str, platform_base_names):
    if not isinstance(plans_str, str):
        return False
    listed = [p.strip() for p in plans_str.split(',') if p.strip()]
    return any(_norm_base_name(p) in platform_base_names for p in listed)


def get_default_markets_for_platform(df_plans: pd.DataFrame, platform: str) -> list[str]:
    mask = df_plans['platform'].fillna("").str.lower() == platform.lower()
    dfp = df_plans[mask]
    markets_set = set()
    for mcell in dfp['markets'].dropna().astype(str):
        markets_set.update(_split_markets_cell(mcell))
    if not markets_set:
        return ['US/CA', 'EUROPE', 'ROW']
    return sorted(markets_set)


def obtener_info_contenido(campaign_name, brief, content_df, plans_df, platform, league_or_other):
    # 0) Copia y columnas de trabajo
    df = content_df.copy()
    df["__name_norm"] = df["content_name"].fillna("").astype(str).apply(_norm_str)
    df["__plans_norm"] = df["plans_available"].fillna("").astype(str)
    df["__plans_norm"] = df["__plans_norm"].apply(lambda s: ", ".join(_norm_base_name(p) for p in s.split(",") if p.strip()))
    league_norm = _norm_str(league_or_other)

    # 1) Filtro por LIGA primero (tolerante)
    if league_norm and league_norm != "otro":
        mask_eq = df["__name_norm"] == league_norm
        mask_contains = df["__name_norm"].str.contains(re.escape(league_norm))
        mask_rev_contains = df["__name_norm"].apply(lambda n: n in league_norm)
        df = df[mask_eq | mask_contains | mask_rev_contains]
        # si a√∫n as√≠ no hay, no salimos todav√≠a; seguimos con filtros por planes

    # 2) Filtro por PLANES v√°lidos de la PLATAFORMA
    platform_base_names, _ = get_platform_plans_set(plans_df, platform)
    # normalizamos tambi√©n el set por si hay "anual"
    platform_base_names = {_norm_base_name(x) for x in platform_base_names}

    def _row_has_any_platform_plan(row):
        row_plans = [p.strip() for p in row.split(",") if p.strip()]
        return any(p in platform_base_names for p in row_plans)

    df = df[df["__plans_norm"].apply(_row_has_any_platform_plan)]

    # 3) Si qued√≥ vac√≠o, intentamos fallback por alias LVBP/Liga Venezolana‚Ä¶
    if df.empty and league_norm and league_norm != "otro":
        aliases = [
            "lvbp",
            "liga venezolana de beisbol profesional",
            "liga venezolana de b√©isbol profesional",
        ]
        df2 = content_df.copy()
        df2["__name_norm"] = df2["content_name"].fillna("").astype(str).apply(_norm_str)
        mask_alias = False
        for a in aliases:
            a_norm = _norm_str(a)
            mask_alias = mask_alias | df2["__name_norm"].str.contains(re.escape(a_norm))
        if mask_alias is not False:
            df = df2[mask_alias]

    # 4) Limpieza de columnas internas
    if df.empty:
        print("ADVERTENCIA: No se encontr√≥ contenido para esta campa√±a (con filtro de plataforma/liga).")
        return pd.DataFrame()
    return df.drop(columns=[c for c in df.columns if c.startswith("__")], errors="ignore")


# ==============================================================================
# 6) OPENAI WRAPPERS
# ==============================================================================

def chat_create(messages, model=None, max_retries=3, timeout=300, **kwargs):
    model = model or MODEL_CHAT
    attempt = 0
    last_err = None
    while attempt <= max_retries:
        try:
            resp = client.chat.completions.create(model=model, messages=messages, timeout=timeout, **kwargs)
            return resp
        except Exception as e:
            last_err = e
            wait = (2 ** attempt) + random.uniform(0, 0.5)
            print(f"[chat_create] Intento {attempt+1}/{max_retries+1} fall√≥: {e}. Reintentando en {wait:.1f}s...")
            sys.stdout.flush()
            time.sleep(wait)
            attempt += 1
    raise last_err


def limpiar_json(texto: str) -> dict:
    print("Intentando limpiar y decodificar la respuesta JSON...")
    try:
        return json.loads(texto)
    except Exception:
        start = texto.find('{'); end = texto.rfind('}') + 1
        if start == -1 or end == 0:
            print("ADVERTENCIA: No se encontr√≥ un objeto JSON v√°lido en la respuesta de la IA.")
            return {}
        json_str = texto[start:end]
        try:
            return json.loads(json_str)
        except Exception as e:
            print(f"ERROR decodificando JSON: {e}\nRespuesta problem√°tica: {json_str[:4000]}")
            return {}


# ==============================================================================
# 7) TEXT BATCH PROCESSORS (SHORTEN/EXPAND/TRANSLATE)
# ==============================================================================

def _expand_batch(texts: list, min_chars: int, max_chars: int, lang: str = 'es') -> tuple:
    idxs = [i for i, t in enumerate(texts) if isinstance(t, str) and t.strip() and len(t.strip()) < min_chars]
    if not idxs:
        return texts, {"prompt_tokens": 0, "completion_tokens": 0, "total_tokens": 0}

    bloques = "\n".join(f"Texto {i+1}: \"{texts[i]}\"" for i in idxs)
    prompts_by_lang = {
        'en': ("You are an expert copywriter who lengthens texts in English while staying under a max character limit.",
               f"Rewrite the following texts so that EACH output has AT LEAST {min_chars} and AT MOST {max_chars} characters. Return only the rewritten texts, one per line, WITHOUT prefixes or quotes."),
        'pt': ("Voc√™ √© um redator especialista que alonga textos em portugu√™s dentro de um limite m√°ximo de caracteres.",
               f"Reescreva para que CADA resultado tenha PELO MENOS {min_chars} e NO M√ÅXIMO {max_chars} caracteres. Retorne apenas os textos reescritos, um por linha, SEM prefixos ou aspas."),
        'es': ("Eres un redactor experto que expande textos en espa√±ol sin superar un l√≠mite de caracteres.",
               f"Reescribe para que CADA salida tenga COMO M√çNIMO {min_chars} y COMO M√ÅXIMO {max_chars} caracteres. Devuelve solo los textos reescritos, uno por l√≠nea, SIN prefijos ni comillas."),
    }
    system_message, prompt_instructions = prompts_by_lang.get(lang, prompts_by_lang['es'])

    resp = chat_create(
        model=MODEL_CHAT,
        messages=[{"role": "system", "content": system_message}, {"role": "user", "content": f"{prompt_instructions}\n\n{bloques}"}],
    )
    usage = getattr(resp, 'usage', None) or {}
    lines = [l.strip() for l in resp.choices[0].message.content.splitlines() if l.strip()]

    cleaned = [re.sub(r'^\s*Texto\s*\d+:\s*"?', '', l).removesuffix('"').strip() for l in lines]
    out = list(texts)
    for i, new in zip(idxs, cleaned):
        out[i] = _smart_trim(new, max_chars)
    return out, usage


def preparar_batch(texts: list, limit: int, tipo: str, lang: str = 'es') -> tuple:
    df = pd.DataFrame({"Original": texts, "Reescrito": texts.copy()})
    total_usage = {"prompt_tokens": 0, "completion_tokens": 0, "total_tokens": 0}

    # 1) Shorten
    idxs_long = df[df["Original"].fillna("").astype(str).str.len() > limit].index.tolist()
    if idxs_long:
        bloques = "\n".join(f'Texto {i+1}: "{df.at[i, "Original"]}"' for i in idxs_long)
        prompts_by_lang = {
            'en': ("You are an expert copywriter who shortens texts in English.", f"Rewrite under {limit} characters. Return only the rewritten texts, one per line, WITHOUT prefixes or quotes."),
            'pt': ("Voc√™ √© um redator especialista que encurta textos em portugu√™s.", f"Reescreva com menos de {limit} caracteres. Retorne um por linha, SEM prefixos nem aspas."),
            'es': ("Eres un redactor experto que acorta textos en espa√±ol.", f"Reescribe con MENOS de {limit} caracteres. Devuelve uno por l√≠nea, SIN prefijos ni comillas."),
        }
        system_message, prompt_instructions = prompts_by_lang.get(lang, prompts_by_lang['es'])
        resp = chat_create(model=MODEL_CHAT, messages=[{"role": "system", "content": system_message}, {"role": "user", "content": f"{prompt_instructions}\n\n{bloques}"}])
        usage = getattr(resp, 'usage', None) or {}
        for k in total_usage:
            total_usage[k] = total_usage.get(k, 0) + (getattr(usage, k, 0) or (usage.get(k, 0) if isinstance(usage, dict) else 0))
        lines = [l.strip() for l in resp.choices[0].message.content.splitlines() if l.strip()]
        cleaned = [re.sub(r'^\s*Texto\s*\d+:\s*"?', '', line).removesuffix('"').strip() for line in lines]
        for i, new_text in zip(idxs_long, cleaned):
            df.at[i, "Reescrito"] = _smart_trim(new_text, limit)

    # 2) Expand if needed
    min_chars = MIN_CHARS_BY_FIELD.get(tipo)
    should_expand = (tipo in MIN_CHARS_BY_FIELD)
    if should_expand and isinstance(min_chars, int) and min_chars > 0:
        current_texts = df["Reescrito"].fillna("").astype(str).tolist()
        needs_expand = any(t.strip() and len(t.strip()) < min_chars for t in current_texts)
        if needs_expand:
            expanded, usage2 = _expand_batch(current_texts, min_chars, limit, lang=lang)
            for k in total_usage:
                total_usage[k] = total_usage.get(k, 0) + (getattr(usage2, k, 0) or 0)
            # Important√≠simo: al final, A√öN si expandi√≥, garantizamos el l√≠mite con smart trim
            df["Reescrito"] = [ _smart_trim(t, limit, lang=lang) for t in expanded ]
    else:
        # Seguridad: nunca expandas headlines / descripciones cortas
        df["Reescrito"] = [ _smart_trim(t, limit, lang=lang) for t in df["Reescrito"].fillna("").astype(str).tolist() ]
        
    return df["Reescrito"].tolist(), total_usage


def traducir_batch(texts: list, target_lang: str) -> tuple:
    if not texts or all(not t for t in texts) or target_lang not in ("en", "pt"):
        return texts, {"prompt_tokens": 0, "completion_tokens": 0, "total_tokens": 0}

    lang_name = "English (US)" if target_lang == "en" else "Portugu√™s (Brasil)"
    block = json.dumps(texts, ensure_ascii=False, indent=2)
    prompt = (
        f"Eres un traductor profesional. Traduce la lista JSON al idioma {lang_name}. "
        f"NO abrevies. Devuelve S√ìLO un objeto JSON con la clave \"translations\".\n{block}"
    )
    try:
        resp = chat_create(
            model=MODEL_CHAT,
            response_format={"type": "json_object"},
            messages=[
                {"role": "system", "content": "You are an expert translator designed to output JSON."},
                {"role": "user", "content": prompt},
            ],
        )
        raw = resp.choices[0].message.content
        return json.loads(raw).get("translations", texts), getattr(resp, 'usage', None)
    except Exception as e:
        print(f"ERROR en traducci√≥n a '{target_lang}': {e}.")
        return texts, {"prompt_tokens": 0, "completion_tokens": 0, "total_tokens": 0}


# ==============================================================================
# 8) PROMPT BUILDER (STRICT JSON & RULES)
# ==============================================================================

def _ejemplos_referencia(ref_df: pd.DataFrame, n=5):
    campo_col = next((c for c in ['campo', 'Campo'] if c in ref_df.columns), None)
    if campo_col:
        df_primary = ref_df[ref_df[campo_col].astype(str).str.lower() == 'primary_texts']
        take_primary = min(2, len(df_primary), n)
        primary_sample = df_primary.sample(n=take_primary, random_state=42) if take_primary > 0 else pd.DataFrame()
        df_rest = ref_df.drop(primary_sample.index)
        remaining = n - len(primary_sample)
        rest_sample = df_rest.sample(n=min(remaining, len(df_rest)), random_state=43) if remaining > 0 else pd.DataFrame()
        ejemplos = pd.concat([primary_sample, rest_sample]).sample(frac=1, random_state=99).reset_index(drop=True)
    else:
        ejemplos = ref_df.sample(n=min(n, len(ref_df)), random_state=42)

    def pick(row, *names): return next((row[n] for n in names if n in row.index), "")
    block = "\n".join(
        f"- [{pick(r,'market','Market')}][{pick(r,'idioma','Idioma')}] {pick(r,'platform','Platform')} {pick(r,'tipo','Tipo')} {pick(r,'campo','Campo')}: \"{pick(r,'texto','Texto')}\""
        for _, r in ejemplos.iterrows()
    )
    return block


def _json_template_for_markets(markets: list[str]) -> dict:
    return {m: {c: {f: ([] if cnt > 1 else "") for f, (cnt, _) in CAMPAIGNS_STRUCTURE.items()} for c in CAMPAIGNS_STRUCTURE} for m in markets}


def _google_pmax_schema_for_market(market: str) -> dict:
    # Solo solicitaremos GooglePMAX para Google; luego replicamos internamente
    fields = CAMPAIGNS_STRUCTURE["GooglePMAX"]
    return {market: {"GooglePMAX": {f: ([] if cnt > 1 else "") for f, (cnt, _) in fields.items()}}}


def _rules_block(plan_info_by_market: dict, base_lang: str) -> str:
    lines = [f"Idioma base de redacci√≥n: {base_lang.upper()}"]
    lines.append("Usa EXCLUSIVAMENTE los planes y precios provistos a continuaci√≥n. Si falta un precio, no lo inventes: deja ese elemento SIN PRECIO y evita mencionarlo.")
    for m, plans in plan_info_by_market.items():
        if not plans:
            lines.append(f"- Mercado {m}: sin planes/valores.")
            continue
        descs = []
        for p in plans:
            # Representaci√≥n plana y clara
            pn = str(p.get('plan_name','') or '').strip()
            cs = str(p.get('currency_symbol','') or '').strip()
            pr = str(p.get('price', '')).strip()
            rp = str(p.get('recurring_period','') or '').strip()
            mkd = _format_discount(p.get('marketing_discount', ''))
            tag = f"{pn} {cs}{pr}/{rp}"
            if mkd:
                tag += f" (PROMO: {mkd})"
            descs.append(tag)
        lines.append(f"- Mercado {m}: {'; '.join(descs)}")
    return "\n".join(lines)


def _google_rules_text():
    return (
        "Reglas Google (estrictas):\n"
        "- PMAX: genera 15 headlines (<=30 car), 5 long_headlines (<=90), 1 short_description (<=60), 4 long_descriptions (<=90).\n"
        "- Demand Generation: se reutilizar√° de PMAX (usar el precio M√ÅS BAJO disponible exacto).\n"
        "- SEM: se reutilizar√° de PMAX (headlines = primeros 14; long_descriptions = mismas 4).\n"
        "- Precios: NUNCA inventar. S√≥lo los provistos.\n"
        "- Se puede destacar EL PLAN ANUAL con promo exacta (si aplica).\n"
    )


def _meta_rules_text():
    return (
        "Reglas Meta (estrictas):\n"
        "- primary_texts: SIEMPRE incluir emojis relacionados al texto (usar los emojis como bullets es una buena pr√°ctica). M√≠nimo 200 caracteres, <=250.\n"
        "- headlines & descriptions: pueden reutilizar titulares/beneficios de Google cuando sean aplicables.\n"
        "- Demand Capture: mencionar expl√≠citamente el plan ANUAL + promo (si existe) con s√≠mbolo y periodo exactos.\n"
        "- NO inventar precios.\n"
    )


def prompt_google_pmax(briefs, ref_df, plan_info_by_market, base_lang='es'):
    ejemplos = _ejemplos_referencia(ref_df, n=5)
    schema = _google_pmax_schema_for_market(briefs['market'])
    rules = _rules_block(plan_info_by_market, base_lang) + "\n" + _google_rules_text()

    prompt = f"""
Eres un generador experto de copies para Google Ads.

Empresa: {briefs['company']}
Campa√±a: {briefs['campaign_name']}
Brief: {briefs['campaign_brief']}
Plataforma objetivo: Google
Mercado: {briefs['market']}

{rules}

Devuelve S√ìLO un JSON con la estructura EXACTA siguiente. NUNCA agregues campos extra ni cambies los nombres.
{json.dumps(schema, ensure_ascii=False, indent=2)}

Ejemplos de copies exitosos (referencia de tono/estilo, no copiar):
{ejemplos}

REGLAS DE ESTILO GLOBALES:
- Estilo: emocional, met√°foras de f√∫tbol, urgencia.
- Moneda: usar EXACTAMENTE el s√≠mbolo indicado en los planes.
- Escribe TODO en Espa√±ol (ES) sin mezclar idiomas.
""".strip()
    return prompt


def prompt_meta_all(briefs, ref_df, plan_info_by_market, base_lang='es'):
    ejemplos = _ejemplos_referencia(ref_df, n=5)

    # Pedimos ambas campa√±as Meta en un solo llamado para consistencia
    template = {
        briefs['market']: {
            "MetaDemandGen": {
                "primary_texts": ["" for _ in range(CAMPAIGNS_STRUCTURE['MetaDemandGen']['primary_texts'][0])],
                "headlines": ["" for _ in range(CAMPAIGNS_STRUCTURE['MetaDemandGen']['headlines'][0])],
                "descriptions": ["" for _ in range(CAMPAIGNS_STRUCTURE['MetaDemandGen']['descriptions'][0])],
            },
            "MetaDemandCapture": {
                "primary_texts": ["" for _ in range(CAMPAIGNS_STRUCTURE['MetaDemandCapture']['primary_texts'][0])],
                "headlines": ["" for _ in range(CAMPAIGNS_STRUCTURE['MetaDemandCapture']['headlines'][0])],
                "descriptions": ["" for _ in range(CAMPAIGNS_STRUCTURE['MetaDemandCapture']['descriptions'][0])],
            }
        }
    }

    rules = _rules_block(plan_info_by_market, base_lang) + "\n" + _meta_rules_text()
    prompt = f"""
Eres un generador experto de copies para Meta Ads.

Empresa: {briefs['company']}
Campa√±a: {briefs['campaign_name']}
Brief: {briefs['campaign_brief']}
Plataforma objetivo: Meta
Mercado: {briefs['market']}

{rules}

Devuelve S√ìLO un JSON con la estructura EXACTA siguiente. NUNCA agregues campos extra ni cambies los nombres.
{json.dumps(template, ensure_ascii=False, indent=2)}

Ejemplos de copies exitosos (referencia de tono/estilo, no copiar):
{ejemplos}

REGLAS DE ESTILO GLOBALES:
- Estilo: emocional, met√°foras de f√∫tbol, urgencia.
- Moneda: usar EXACTAMENTE el s√≠mbolo indicado en los planes.
- Escribe TODO en Espa√±ol (ES) sin mezclar idiomas.
""".strip()
    return prompt


# ==============================================================================
# 9) VALIDATION & ENFORCEMENT HELPERS
# ==============================================================================

def _ensure_list_of_len(value, length):
    arr = value if isinstance(value, list) else ([value] if isinstance(value, str) else [])
    arr = [s if isinstance(s, str) else "" for s in arr]
    if len(arr) < length:
        arr += [""] * (length - len(arr))
    return arr[:length]


def _trim_to_limits(arr, limit):
    return [_smart_trim(s or "", limit) for s in arr]


def _post_enforce_counts_and_limits(struct: dict):
    # Recorre markets/campaigns/fields y ajusta a conteos & l√≠mites
    for market, cmap in struct.items():
        for campaign, fields in CAMPAIGNS_STRUCTURE.items():
            if campaign not in cmap:
                continue
            for field, (count, limit) in fields.items():
                if field in cmap[campaign]:
                    cmap[campaign][field] = _ensure_list_of_len(cmap[campaign][field], count)
                    cmap[campaign][field] = _trim_to_limits(cmap[campaign][field], limit)
    return struct


def _lowest_price_token(plans: list[dict]) -> str | None:
    # Encuentra el menor precio disponible (independiente de periodo), devuelve string listo "$9.99"
    best = None
    for p in plans or []:
        try:
            price = float(str(p.get('price', '')).replace(',', '.'))
            sym = (p.get('currency_symbol') or '').strip()
            if math.isnan(price):
                continue
            token = f"{sym}{price:.2f}".rstrip('0').rstrip('.')
            if (best is None) or (price < best[0]):
                best = (price, token)
        except Exception:
            continue
    return best[1] if best else None


def _annual_plan_with_promo(plans: list[dict]) -> dict | None:
    for p in plans or []:
        rp = (p.get('recurring_period') or '').lower()
        if 'annual' in rp or 'anual' in rp:
            if p.get('marketing_discount'):
                return p
    return None


def _replicate_google_from_pmax(google_struct_for_market: dict):
    # Assumes structure: {market: {GooglePMAX:{...}}}
    market = next(iter(google_struct_for_market.keys()))
    pmax = google_struct_for_market[market].get('GooglePMAX', {})

    # Prepare copies
    sem = {"headlines": [], "long_headlines": []}
    ddg = {"headlines": [], "short_description": [], "long_descriptions": []}

    sem_head_count = CAMPAIGNS_STRUCTURE['SEM']['headlines'][0]
    sem_long_count = CAMPAIGNS_STRUCTURE['SEM']['long_headlines'][0]
    ddg_heads = CAMPAIGNS_STRUCTURE['GoogleDemandGen']['headlines'][0]

    sem['headlines'] = _ensure_list_of_len(pmax.get('headlines', []), CAMPAIGNS_STRUCTURE['GooglePMAX']['headlines'][0])[:sem_head_count]
    sem['long_headlines'] = _ensure_list_of_len(pmax.get('long_headlines', []), CAMPAIGNS_STRUCTURE['GooglePMAX']['long_headlines'][0])[:sem_long_count]

    ddg['headlines'] = _ensure_list_of_len(pmax.get('headlines', []), CAMPAIGNS_STRUCTURE['GooglePMAX']['headlines'][0])[:ddg_heads]
    ddg['short_description'] = _ensure_list_of_len(pmax.get('short_description', []), 1)
    ddg['long_descriptions'] = _ensure_list_of_len(pmax.get('long_descriptions', []), CAMPAIGNS_STRUCTURE['GooglePMAX']['long_descriptions'][0])[:CAMPAIGNS_STRUCTURE['GoogleDemandGen']['long_descriptions'][0]]

    google_struct_for_market[market]['SEM'] = sem
    google_struct_for_market[market]['GoogleDemandGen'] = ddg

    # Enforce limits
    for camp in ['SEM', 'GoogleDemandGen', 'GooglePMAX']:
        if camp not in google_struct_for_market[market]:
            continue
        for field, (count, limit) in CAMPAIGNS_STRUCTURE[camp].items():
            if field in google_struct_for_market[market][camp]:
                google_struct_for_market[market][camp][field] = _ensure_list_of_len(google_struct_for_market[market][camp][field], count)
                google_struct_for_market[market][camp][field] = _trim_to_limits(google_struct_for_market[market][camp][field], limit)

    return google_struct_for_market


# ==============================================================================
# 10) EXCEL WRITER
# ==============================================================================

def generar_excel_multi(data: dict, output_langs: tuple = ("es",), filename: str = "copies.xlsx") -> dict:
    all_tasks = []
    total_usage = {"prompt_tokens": 0, "completion_tokens": 0}

    # Fase 1 (ES): enforce counts/limits beforehand
    for market, cmap in data.items():
        for campaign, fields in CAMPAIGNS_STRUCTURE.items():
            plat = {'SEM': 'Google', 'GoogleDemandGen': 'Google', 'GooglePMAX': 'Google',
                    'MetaDemandGen': 'Meta', 'MetaDemandCapture': 'Meta'}[campaign]
            tp = {'SEM': 'SEM', 'GoogleDemandGen': 'DemandGen', 'GooglePMAX': 'PMAX',
                  'MetaDemandGen': 'DemandGen', 'MetaDemandCapture': 'DemandCapture'}[campaign]
            campaign_data = cmap.get(campaign, {})
            for field, (count, limit) in fields.items():
                original = campaign_data.get(field, [])
                if not isinstance(original, list):
                    original = [original]
                original = _ensure_list_of_len(original, count)

                # Fix English, expand primary_texts, etc.
                es_texts, usage = preparar_batch(original, limit, field, lang='es')
                if any(_seems_english(t) for t in es_texts):
                    es_texts, usage_fix = preparar_batch(es_texts, limit, field, lang='es')
                    for k in total_usage:
                        total_usage[k] += getattr(usage_fix, k, 0) or 0

                for k in total_usage:
                    total_usage[k] += getattr(usage, k, 0) or 0

                all_tasks.append({
                    "market": market, "platform": plat, "tipo": tp, "campo": field,
                    "count": count, "limit": limit, "es_texts": es_texts,
                })

    # Fase 2: Traducciones
    need_en = "en" in output_langs
    need_pt = "pt" in output_langs

    rows = []
    for task in all_tasks:
        es_texts, limit, campo = task['es_texts'], task['limit'], task['campo']
        en_texts, pt_texts = [], []
        has_content = any(t.strip() for t in es_texts)

        if need_en and has_content:
            en_raw, en_u1 = traducir_batch(es_texts, 'en')
            en_texts, en_u2 = preparar_batch(en_raw, limit, campo, lang='en')
            for k in total_usage:
                total_usage[k] += (getattr(en_u1, k, 0) or (en_u1.get(k,0) if isinstance(en_u1, dict) else 0))
                total_usage[k] += (getattr(en_u2, k, 0) or (en_u2.get(k,0) if isinstance(en_u2, dict) else 0))
        if need_pt and has_content:
            pt_raw, pt_u1 = traducir_batch(es_texts, 'pt')
            pt_texts, pt_u2 = preparar_batch(pt_raw, limit, campo, lang='pt')
            for k in total_usage:
                total_usage[k] += (getattr(pt_u1, k, 0) or (pt_u1.get(k,0) if isinstance(pt_u1, dict) else 0))
                total_usage[k] += (getattr(pt_u2, k, 0) or (pt_u2.get(k,0) if isinstance(pt_u2, dict) else 0))

        langs_map = {'es': es_texts}
        if need_en: langs_map['en'] = en_texts
        if need_pt: langs_map['pt'] = pt_texts

        for lang, texts in langs_map.items():
            for i in range(task['count']):
                txt = texts[i] if i < len(texts) else ""
                if txt is None:
                    txt = ""
                rows.append({
                    "Market": task['market'], "Platform": task['platform'], "Tipo": task['tipo'],
                    "Campo": task['campo'], "T√≠tulo": f"{task['campo']} {i+1}", "Idioma": lang,
                    "Texto": txt, "Caracteres": len(txt), "Max Caracteres": limit,
                    "Check": 1 if len(txt) <= limit else 0,
                })

    df_master = pd.DataFrame(rows)
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df_master.to_excel(writer, sheet_name='Todos los Copies', index=False)
        for (platform, tipo), df_group in df_master.groupby(['Platform', 'Tipo']):
            df_group.sort_values(by=['Market', 'Idioma', 'Campo']).to_excel(writer, sheet_name=f"{platform} {tipo}", index=False)

    wb = load_workbook(filename)
    red_font = Font(color="9C0006")
    red_fill = PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE")
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for col in ws.columns:
            column_letter = col[0].column_letter
            if ws[f"{column_letter}1"].value == 'Texto':
                ws.column_dimensions[column_letter].width = 60
                continue
            max_length = max(len(str(cell.value)) for cell in col if cell.value is not None)
            ws.column_dimensions[column_letter].width = min(80, max_length + 2)
        if sheet_name == 'Todos los Copies':
            for row in ws.iter_rows(min_row=2):
                if row[9].value == 0:  # 'Check'
                    for cell in row:
                        cell.font = red_font
                        cell.fill = red_fill
    wb.save(filename)
    return total_usage


# ==============================================================================
# 11) ORCHESTRATION (PLATFORM SPLIT & PMAX REUSE)
# ==============================================================================

def generar_copies(
    campaign_name: str,
    campaign_brief: str,
    platform_name: str = "Fanatiz",
    langs_csv: str = "ES",
    league_selection: str = "Otro",
    output_filename: str = "copies_generadas.xlsx",
    markets_selected: list[str] | None = None,
) -> tuple:
    print(f"Iniciando generaci√≥n para '{campaign_name}' | Plataforma: {platform_name} | Liga: {league_selection} | Langs: {langs_csv}")

    # 1) Load data
    total_usage = {"prompt_tokens": 0, "completion_tokens": 0}
    base_path = os.path.abspath(os.path.dirname(__file__))
    df_refs = _normalize_headers(cargar_referencias(os.path.join(base_path, "Mejor_Performing_Copies_Paid_Fanatiz.xlsx")))
    df_content = _normalize_headers(cargar_contenidos(os.path.join(base_path, "content_by_country.xlsx")))
    df_plans = _normalize_headers(cargar_planes(os.path.join(base_path, "plans_and_pricing.xlsx")))
    df_specs = _normalize_headers(cargar_specs(os.path.join(base_path, "platforms_and_campaigns_specs.xlsx")))

    _ensure_columns(df_plans, ['platform', 'plan_name', 'markets', 'recurring_period'], 'plans_and_pricing')
    _ensure_columns(df_content, ['content_name', 'markets_available', 'plans_available', 'content_languages'], 'content_by_country')

    # 2) Find content
    relevant = obtener_info_contenido(campaign_name, campaign_brief, df_content, df_plans, platform_name, league_selection)

    # Fallback
    if relevant.empty or (league_selection and league_selection.lower() == "otro"):
        print(f"\n‚ö†Ô∏è Fallback a contenido por defecto para la plataforma '{platform_name}' (liga '{league_selection}')")
        relevant = pd.DataFrame({
            'content_name': [campaign_name],
            'markets_available': [",".join(get_default_markets_for_platform(df_plans, platform_name))],
            'plans_available': [",".join(DEFAULT_PLANS_BY_PLATFORM.get(platform_name, []))],
            'content_languages': [langs_csv.upper()],
            'content_details': [f'Contenido general de {platform_name}'],
        })

    # 3) Markets
    all_markets = {item.strip() for m_list in relevant['markets_available'].dropna().str.split(',') for item in m_list if item.strip()}
    platform_markets = set(get_default_markets_for_platform(df_plans, platform_name))
    markets_to_process = sorted(list(all_markets & platform_markets)) if platform_markets else sorted(list(all_markets))
    if markets_selected:
        sel = {m.strip().upper() for m in markets_selected}
        markets_to_process = [m for m in markets_to_process if m.upper() in sel]
    if not markets_to_process:
        msg = "Error: no se encontraron mercados v√°lidos para procesar tras aplicar los filtros."
        print(f"‚ö†Ô∏è {msg}")
        return output_filename, msg

    # 4) Generate per platform split
    final_by_market = {}
    platform_base_names, _ = get_platform_plans_set(df_plans, platform_name)

    for market in markets_to_process:
        print("\n" + "="*20 + f" PROCESANDO MERCADO: {market} " + "="*20)

        market_df = relevant[relevant['markets_available'].str.contains(market, na=False, case=False)]
        plans_in_content = {p.strip() for p_list in market_df['plans_available'].dropna().str.split(',') for p in p_list if p.strip()}
        valid_plans = {p for p in plans_in_content if _norm_base_name(p) in platform_base_names}
        if not valid_plans:
            print(f"  ‚ö†Ô∏è No se encontraron planes v√°lidos para '{platform_name}' en '{market}'. Saltando...")
            continue

        # Collect plan detailed rows for this market
        plan_info_market = []
        for plan_name in sorted(list(valid_plans)):
            mask = (
                (df_plans["platform"].fillna("").str.lower() == platform_name.lower()) &
                (df_plans["plan_name"].str.contains(plan_name.split()[0], case=False, na=False)) &
                (df_plans["markets"].str.contains(market, case=False, na=False))
            )
            if "annual" in plan_name.lower() or "anual" in plan_name.lower():
                mask &= (df_plans["recurring_period"].str.lower().isin(["annual", "anual"]))
            if "monthly" in plan_name.lower():
                mask &= (df_plans["recurring_period"].str.lower() == "monthly")
            sel = df_plans[mask]
            if not sel.empty:
                plan_info_market.append(sel.iloc[0].to_dict())

        briefs = {
            'campaign_name': campaign_name,
            'campaign_brief': campaign_brief,
            'company': COMPANY_BY_PLATFORM.get(platform_name, platform_name),
            'extras': f'Plataforma objetivo: {platform_name}',
            'market': market,
        }

        # ===== Google (solo PMAX, luego replicamos) =====
        google_data_market = {}
        try:
            prompt_g = prompt_google_pmax(briefs, df_refs, {market: plan_info_market}, base_lang='es')
            resp_g = chat_create(
                model=MODEL_CHAT,
                response_format={"type": "json_object"},
                messages=[
                    {"role": "system", "content": "You are a helpful assistant that outputs STRICT JSON only."},
                    {"role": "user", "content": prompt_g},
                ],
            )
            total_usage["prompt_tokens"] += getattr(resp_g.usage, 'prompt_tokens', 0) or 0
            total_usage["completion_tokens"] += getattr(resp_g.usage, 'completion_tokens', 0) or 0
            market_data_g = limpiar_json(resp_g.choices[0].message.content) or {}
        except Exception as e:
            print(f"Error en generaci√≥n Google PMAX para {market}: {e}")
            market_data_g = {}

        # Asegurar claves y replicar
        if market not in market_data_g:
            market_data_g[market] = {"GooglePMAX": {f: [] for f in CAMPAIGNS_STRUCTURE['GooglePMAX'].keys()}}
        google_data_market = _replicate_google_from_pmax(copy.deepcopy(market_data_g))

        # ===== Meta (DemandGen & DemandCapture) =====
        meta_data_market = {}
        try:
            prompt_m = prompt_meta_all(briefs, df_refs, {market: plan_info_market}, base_lang='es')
            resp_m = chat_create(
                model=MODEL_CHAT,
                response_format={"type": "json_object"},
                messages=[
                    {"role": "system", "content": "You are a helpful assistant that outputs STRICT JSON only."},
                    {"role": "user", "content": prompt_m},
                ],
            )
            total_usage["prompt_tokens"] += getattr(resp_m.usage, 'prompt_tokens', 0) or 0
            total_usage["completion_tokens"] += getattr(resp_m.usage, 'completion_tokens', 0) or 0
            meta_data_market = limpiar_json(resp_m.choices[0].message.content) or {}
        except Exception as e:
            print(f"Error en generaci√≥n Meta para {market}: {e}")
            meta_data_market = {}

        # Merge Google + Meta
        merged_market = {market: {}}
        merged_market[market].update(google_data_market.get(market, {}))
        merged_market[market].update(meta_data_market.get(market, {}))

        # Post-enforce counts/limits
        merged_market = _post_enforce_counts_and_limits(merged_market)

        # Reglas sem√°nticas extra: ajustar menciones de precio
        # DemandGen (Google/Meta): usar menor precio si se menciona precio
        low_token = _lowest_price_token(plan_info_market)
        if low_token:
            for camp in ["GoogleDemandGen", "MetaDemandGen"]:
                if camp in merged_market[market]:
                    for field in merged_market[market][camp]:
                        arr = merged_market[market][camp][field]
                        merged_market[market][camp][field] = [re.sub(r"\bdesde\b.*?(\d+[\.,]?\d*)?", low_token, s, flags=re.IGNORECASE) for s in arr]
        # DemandCapture (Meta): destacar ANUAL con promo
        anual = _annual_plan_with_promo(plan_info_market)
        if anual:
            sym   = str(anual.get('currency_symbol', '') or '').strip()
            price = str(anual.get('price', '') or '').strip()
            rp    = str(anual.get('recurring_period', '') or '').strip()
            mkd   = _format_discount(anual.get('marketing_discount', ''))
            tag = f"{sym}{price}/{rp}" + (f" ({mkd} de descuento)" if mkd else "")
            if 'MetaDemandCapture' in merged_market[market]:
                for field in merged_market[market]['MetaDemandCapture']:
                    merged_market[market]['MetaDemandCapture'][field] = [
                        re.sub(r"\bplan anual\b(?![^)]*\))", f"plan anual {tag}", s, flags=re.IGNORECASE) if isinstance(s, str) else s
                        for s in merged_market[market]['MetaDemandCapture'][field]
                    ]

        # Save
        final_by_market[market] = merged_market[market]

    if not final_by_market:
        return output_filename, "Error: No se generaron copies v√°lidos despu√©s de procesar todos los mercados."

    # 5) Excel
    excel_usage = generar_excel_multi(final_by_market, output_langs=_parse_langs(langs_csv), filename=output_filename)
    for k in total_usage:
        total_usage[k] += excel_usage.get(k, 0) or 0

    # Costos estimados
    PRICE_PER_MILLION_INPUT = float(os.getenv("PRICE_PM_INPUT", 0.250))
    PRICE_PER_MILLION_OUTPUT = float(os.getenv("PRICE_PM_OUTPUT", 2.000))
    input_cost = (total_usage.get("prompt_tokens", 0) / 1_000_000) * PRICE_PER_MILLION_INPUT
    output_cost = (total_usage.get("completion_tokens", 0) / 1_000_000) * PRICE_PER_MILLION_OUTPUT
    total_cost = input_cost + output_cost

    summary = (
        f"üßæ **Campa√±a:** {campaign_name}\n"
        f"üìù **Brief:** {campaign_brief}\n\n"
        f"üìä **Resumen de Consumo y Costo** üí∞\n"
        f"-----------------------------------------\n"
        f"Tokens Entrada: {total_usage.get('prompt_tokens', 0):,}\n"
        f"Tokens Salida:  {total_usage.get('completion_tokens', 0):,}\n"
        f"**Tokens Totales:** {total_usage.get('prompt_tokens', 0) + total_usage.get('completion_tokens', 0):,}\n"
        f"Costo Total Estimado: **${total_cost:.4f} USD**\n"
    )
    print(summary)
    print(f"¬°Proceso completado! Archivo guardado en: {output_filename}")
    return output_filename, summary




